import os
import re
import asyncio
from datetime import datetime, timezone, timedelta
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
from playwright.async_api import async_playwright

URL = "https://fund.ivyro.net/dataroom/om_data/om_dataV2_C2B9_B300_C218.htm"
BASE = Path(__file__).resolve().parent
EXCEL_PATH = BASE / "samo.xlsx"
CAPTURE_DIR = BASE / "captures"
CHART_DIR = BASE / "charts"
CAPTURE_DIR.mkdir(exist_ok=True)
CHART_DIR.mkdir(exist_ok=True)
KST = timezone(timedelta(hours=9))


def decode_page(resp: requests.Response) -> str:
    for enc in ("cp949", "euc-kr", "utf-8"):
        try:
            return resp.content.decode(enc)
        except UnicodeDecodeError:
            pass
    return resp.text


def get_text_and_values():
    r = requests.get(URL, timeout=20)
    r.raise_for_status()
    html = decode_page(r)
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(" ", strip=True)
    nums = [m.group(0).replace(",", "") for m in re.finditer(r"\d{1,3}(?:,\d{3})+|\d+(?:\.\d+)?", text)]
    print("Number count:", len(nums))

    if len(nums) < 19:
        (BASE / "debug_text.txt").write_text(text, encoding="utf-8")
        raise RuntimeError("Not enough numbers extracted. debug_text.txt saved.")

    principal_raw = int(float(nums[11]))
    daily_profit = int(float(nums[16]))
    monthly_profit = int(float(nums[17]))
    total_profit = int(float(nums[18]))
    principal = round(principal_raw / 10000)
    return text, daily_profit, monthly_profit, total_profit, principal, principal_raw


def page_date_matches(text: str, today: datetime) -> bool:
    y, m, d = today.year, today.month, today.day
    patterns = [
        rf"{y}\D{{0,10}}{m}\D{{0,10}}{d}",
        rf"{y}-{m:02d}-{d:02d}",
    ]
    return any(re.search(p, text) for p in patterns)


async def capture_page(crop_path: Path):
    full_path = crop_path.with_name("full_" + crop_path.name)
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        page = await browser.new_page(viewport={"width": 1920, "height": 3000})
        await page.goto(URL, wait_until="networkidle", timeout=60000)
        await page.screenshot(path=str(full_path), full_page=True)
        await browser.close()

    img = Image.open(full_path)
    crop = img.crop((0, 0, 1050, 620))
    crop.save(crop_path)
    full_path.unlink(missing_ok=True)


def ensure_workbook():
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        wb.active.title = "DailyProfit"

    # 한글 시트가 있으면 사용, 없으면 영어 시트 사용
    if "일일수익" in wb.sheetnames:
        ws = wb["일일수익"]
    elif "DailyProfit" in wb.sheetnames:
        ws = wb["DailyProfit"]
    else:
        ws = wb.active
        ws.title = "DailyProfit"

    headers = [
        "Date", "DailyProfit", "MonthlyProfit", "TotalProfit", "Principal",
        "DailyRate", "MonthlyRate", "TotalRate", "DailyChange", "TotalChange"
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i).value = h
    return wb, ws


def to_int(value, default=0):
    if value is None:
        return default
    try:
        return int(float(str(value).replace(",", "").replace("%", "").strip()))
    except Exception:
        return default


def update_excel(today_str, daily, monthly, total, principal, principal_raw):
    wb, ws = ensure_workbook()

    prev_daily = 0
    prev_total = 0
    target_row = None

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, 1).value
        if not cell:
            continue

        # 오늘 날짜 행이 있으면 업데이트 대상으로 지정
        if str(cell)[:10] == today_str:
            target_row = row
            continue

        # 숫자인 데이터 행만 전일 값 후보로 사용. 헤더/문자열 행은 자동 무시.
        candidate_daily = to_int(ws.cell(row, 2).value, None)
        candidate_total = to_int(ws.cell(row, 4).value, None)
        if candidate_daily is not None:
            prev_daily = candidate_daily
        if candidate_total is not None:
            prev_total = candidate_total

    if target_row is None:
        target_row = ws.max_row + 1

    daily_rate = daily / principal_raw if principal_raw else 0
    monthly_rate = monthly / principal_raw if principal_raw else 0
    total_rate = total / principal_raw if principal_raw else 0
    daily_change = daily - prev_daily
    total_change = total - prev_total

    values = [
        today_str,
        daily,
        monthly,
        total,
        principal,
        daily_rate,
        monthly_rate,
        total_rate,
        daily_change,
        total_change,
    ]

    for col, val in enumerate(values, start=1):
        ws.cell(target_row, col).value = val

    for col in range(2, 5):
        ws.cell(target_row, col).number_format = "#,##0"
    ws.cell(target_row, 5).number_format = "#,##0"
    for col in range(6, 9):
        ws.cell(target_row, col).number_format = "0.00%"
    for col in range(9, 11):
        ws.cell(target_row, col).number_format = "#,##0"
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 14

    update_stats_sheet(wb, ws)
    update_excel_charts(wb, ws)
    wb.save(EXCEL_PATH)
    return daily_change, total_change, daily_rate, monthly_rate, total_rate


def sheet_rows(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        date_value = ws.cell(r, 1).value
        if not date_value:
            continue
        daily = to_int(ws.cell(r, 2).value, None)
        total = to_int(ws.cell(r, 4).value, None)
        if daily is None or total is None:
            continue
        rows.append({
            "date": str(date_value)[:10],
            "daily": float(daily),
            "monthly": float(to_int(ws.cell(r, 3).value, 0)),
            "total": float(total),
            "principal": float(to_int(ws.cell(r, 5).value, 0)),
            "daily_rate": float(ws.cell(r, 6).value or 0),
            "monthly_rate": float(ws.cell(r, 7).value or 0),
            "total_rate": float(ws.cell(r, 8).value or 0),
        })
    return rows


def update_stats_sheet(wb, ws):
    if "Stats" in wb.sheetnames:
        del wb["Stats"]
    st = wb.create_sheet("Stats")
    rows = sheet_rows(ws)
    if not rows:
        return

    daily_values = [r["daily"] for r in rows]
    win_count = len([v for v in daily_values if v > 0])
    avg_daily = sum(daily_values) / len(daily_values)
    best = max(rows, key=lambda r: r["daily"])
    worst = min(rows, key=lambda r: r["daily"])
    latest = rows[-1]

    peak = rows[0]["total"]
    mdd = 0
    for r in rows:
        peak = max(peak, r["total"])
        if peak:
            mdd = min(mdd, (r["total"] - peak) / peak)

    stats = [
        ("Current Date", latest["date"]),
        ("Principal", latest["principal"]),
        ("Daily Profit", latest["daily"]),
        ("Monthly Profit", latest["monthly"]),
        ("Total Profit", latest["total"]),
        ("Daily Rate", latest["daily_rate"]),
        ("Monthly Rate", latest["monthly_rate"]),
        ("Total Rate", latest["total_rate"]),
        ("Average Daily Profit", avg_daily),
        ("Best Day", best["date"]),
        ("Best Daily Profit", best["daily"]),
        ("Worst Day", worst["date"]),
        ("Worst Daily Profit", worst["daily"]),
        ("Win Rate", win_count / len(daily_values)),
        ("MDD", mdd),
    ]

    st.append(["Metric", "Value"])
    for k, v in stats:
        st.append([k, v])
    st.column_dimensions["A"].width = 24
    st.column_dimensions["B"].width = 18
    for row in range(2, st.max_row + 1):
        metric = st.cell(row, 1).value
        if metric in ("Daily Rate", "Monthly Rate", "Total Rate", "Win Rate", "MDD"):
            st.cell(row, 2).number_format = "0.00%"
        elif isinstance(st.cell(row, 2).value, (int, float)):
            st.cell(row, 2).number_format = "#,##0"


def update_excel_charts(wb, ws):
    if "Charts" in wb.sheetnames:
        del wb["Charts"]
    chs = wb.create_sheet("Charts")
    max_row = ws.max_row
    if max_row < 3:
        return

    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)

    line = LineChart()
    line.title = "Daily Profit"
    line.y_axis.title = "Profit"
    line.x_axis.title = "Date"
    data = Reference(ws, min_col=2, min_row=1, max_row=max_row)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    chs.add_chart(line, "A1")

    total_chart = LineChart()
    total_chart.title = "Total Profit"
    data = Reference(ws, min_col=4, min_row=1, max_row=max_row)
    total_chart.add_data(data, titles_from_data=True)
    total_chart.set_categories(cats)
    chs.add_chart(total_chart, "A18")

    rate = LineChart()
    rate.title = "Return Rates"
    data = Reference(ws, min_col=6, max_col=8, min_row=1, max_row=max_row)
    rate.add_data(data, titles_from_data=True)
    rate.set_categories(cats)
    chs.add_chart(rate, "A35")


def make_png_charts():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["일일수익"] if "일일수익" in wb.sheetnames else wb["DailyProfit"]
    rows = sheet_rows(ws)
    if not rows:
        return []

    dates = [r["date"][5:] for r in rows]
    outputs = []

    def save_line(y, title, ylabel, filename):
        path = CHART_DIR / filename
        plt.figure(figsize=(10, 5))
        plt.plot(dates, y, marker="o")
        plt.title(title)
        plt.xlabel("Date")
        plt.ylabel(ylabel)
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.savefig(path, dpi=150)
        plt.close()
        outputs.append(path)

    save_line([r["daily"] for r in rows], "Daily Profit", "KRW", "chart_daily_profit.png")
    save_line([r["monthly"] for r in rows], "Monthly Profit", "KRW", "chart_monthly_profit.png")
    save_line([r["total"] for r in rows], "Total Profit", "KRW", "chart_total_profit.png")
    save_line([r["total_rate"] * 100 for r in rows], "Total Return Rate", "%", "chart_total_rate.png")
    return outputs


def telegram_send_message(text):
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        print("Telegram secrets not set. Skip message.")
        return
    requests.post(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data={"chat_id": chat_id, "text": text},
        timeout=20,
    ).raise_for_status()


def telegram_send_photo(path: Path, caption: str):
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        print("Telegram secrets not set. Skip photo.")
        return
    with open(path, "rb") as f:
        requests.post(
            f"https://api.telegram.org/bot{token}/sendPhoto",
            data={"chat_id": chat_id, "caption": caption},
            files={"photo": f},
            timeout=60,
        ).raise_for_status()


def fmt(n):
    return f"{int(round(n)):,}"


def fmt_change(n):
    n = int(round(n))
    if n > 0:
        return f"+{n:,}"
    return f"{n:,}"


def pct(x):
    return f"{x * 100:.2f}%"


async def main():
    today = datetime.now(KST)
    today_str = today.strftime("%Y-%m-%d")
    text, daily, monthly, total, principal, principal_raw = get_text_and_values()

    force = os.environ.get("FORCE_RUN", "false").lower() == "true"
    if not force and not page_date_matches(text, today):
        msg = f"[Samo Fund Skip]\nDate not matched. Today: {today_str}"
        print(msg)
        telegram_send_message(msg)
        return

    print("DailyProfit:", daily)
    print("MonthlyProfit:", monthly)
    print("TotalProfit:", total)
    print("Principal:", principal)

    capture_path = CAPTURE_DIR / f"fund_capture_{today.strftime('%Y%m%d_%H%M%S')}.png"
    await capture_page(capture_path)

    daily_change, total_change, daily_rate, monthly_rate, total_rate = update_excel(
        today_str, daily, monthly, total, principal, principal_raw
    )
    chart_paths = make_png_charts()

    message = (
        "[Samo Fund Capture Completed]\n\n"
        f"Date : {today_str}\n"
        f"Daily Profit : {fmt(daily)}\n"
        f"Daily Change : {fmt_change(daily_change)}\n"
        f"Monthly Profit : {fmt(monthly)}\n"
        f"Total Profit : {fmt(total)}\n"
        f"Total Change : {fmt_change(total_change)}\n"
        f"Principal : {fmt(principal)}\n\n"
        f"Daily Rate : {pct(daily_rate)}\n"
        f"Monthly Rate : {pct(monthly_rate)}\n"
        f"Total Rate : {pct(total_rate)}"
    )

    telegram_send_message(message)
    telegram_send_photo(capture_path, "Fund Capture")
    for p in chart_paths:
        telegram_send_photo(p, p.stem)
    print("Done")


if __name__ == "__main__":
    asyncio.run(main())
